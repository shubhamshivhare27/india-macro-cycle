# ============================================================
#  dashboard/streamlit_dashboard.py
#  Run with: streamlit run dashboard/streamlit_dashboard.py
#  Shows live cycle phase, indicator table, ETF recs, history
# ============================================================

import sys
import json
import csv
import datetime
from pathlib import Path

# Allow imports from parent
sys.path.insert(0, str(Path(__file__).parent.parent))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px

from config import (DATA_DIR, EXCEL_PATH, EXCEL_SHEET, SECTOR_STRATEGY,
                    INDICATOR_META)

# ── Page config ───────────────────────────────────────────────
st.set_page_config(
    page_title="India Macro Cycle Dashboard",
    page_icon="🇮🇳",
    layout="wide",
    initial_sidebar_state="expanded",
)

PHASE_COLOURS = {
    "STRONG RECOVERY": "#1a7f37",
    "EARLY EXPANSION": "#2d8b56",
    "MID CYCLE":       "#b7950b",
    "LATE CYCLE":      "#8e44ad",
    "CONTRACTION":     "#c0392b",
    "UNKNOWN":         "#95a5a6",
}

# ── Load data ─────────────────────────────────────────────────
@st.cache_data(ttl=300)
def load_snapshot():
    snap_path = DATA_DIR / "macro_snapshot.json"
    if snap_path.exists():
        with open(snap_path) as f:
            return json.load(f)
    return {}

@st.cache_data(ttl=300)
def load_history():
    hist_path = DATA_DIR / "macro_history.csv"
    if hist_path.exists():
        df = pd.read_csv(hist_path)
        df["date"] = pd.to_datetime(df["date"])
        return df
    return pd.DataFrame()

def load_cycle_result():
    snap = load_snapshot()
    return snap.get("cycle_result", {})

def read_excel_score():
    """Read current score directly from Excel."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(EXCEL_PATH), data_only=True)
        ws = wb[EXCEL_SHEET]
        score = float(ws["E39"].value or 0)
        label = str(ws["H49"].value or "")
        wb.close()
        return score, label
    except Exception:
        return None, None

# ── Sidebar ───────────────────────────────────────────────────
st.sidebar.image("https://upload.wikimedia.org/wikipedia/en/4/41/Flag_of_India.svg", width=80)
st.sidebar.title("India Macro Cycle")
st.sidebar.caption(f"Last refreshed: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}")

if st.sidebar.button("🔄 Refresh Data"):
    st.cache_data.clear()
    st.rerun()

st.sidebar.markdown("---")
st.sidebar.markdown("**Quick Links**")
st.sidebar.markdown("- [RBI](https://www.rbi.org.in)")
st.sidebar.markdown("- [MOSPI](https://mospi.gov.in)")
st.sidebar.markdown("- [Finance Min GST](https://www.gst.gov.in)")
st.sidebar.markdown("- [SIAM Auto Data](https://www.siam.in)")
st.sidebar.markdown("- [NSE India](https://www.nseindia.com)")

# ── Main ──────────────────────────────────────────────────────
st.title("🇮🇳 India Business Cycle Dashboard")
st.caption("Merrill Lynch Investment Clock — adapted for India | Free official sources")

snapshot  = load_snapshot()
hist_df   = load_history()
indicators = snapshot.get("indicators", {})

# Try reading directly from Excel
excel_score, excel_label = read_excel_score()

# Determine current score and phase
if excel_score is not None:
    composite_score = excel_score
    raw_label       = excel_label or ""
else:
    composite_score = snapshot.get("cycle_result", {}).get("composite_score", 0.0)
    raw_label       = snapshot.get("cycle_result", {}).get("phase", "UNKNOWN")

# Derive phase
def score_to_phase(s):
    if s >= 0.80: return "STRONG RECOVERY"
    if s >= 0.60: return "EARLY EXPANSION"
    if s >= 0.40: return "MID CYCLE"
    if s >= 0.20: return "LATE CYCLE"
    return "CONTRACTION"

phase = score_to_phase(composite_score)
phase_colour = PHASE_COLOURS.get(phase, "#555")

# ── Row 1: Key Metrics ────────────────────────────────────────
c1, c2, c3, c4, c5 = st.columns(5)

with c1:
    st.markdown(f"""<div style="background:{phase_colour};color:white;padding:20px;
                    border-radius:10px;text-align:center;">
        <div style="font-size:13px;">CYCLE PHASE</div>
        <div style="font-size:20px;font-weight:bold;">{phase}</div>
    </div>""", unsafe_allow_html=True)

with c2:
    st.metric("Composite Score", f"{composite_score:.4f}",
              delta=f"{snapshot.get('cycle_result',{}).get('score_delta',0) or 0:+.4f}")

with c3:
    momentum = snapshot.get("cycle_result",{}).get("momentum","—")
    st.metric("Momentum", momentum)

with c4:
    conf = snapshot.get("cycle_result",{}).get("confidence","—")
    live = snapshot.get("cycle_result",{}).get("live_count",0)
    st.metric("Confidence", conf, delta=f"{live}/15 live")

with c5:
    rebalance = snapshot.get("cycle_result",{}).get("rebalance_flag", False)
    st.metric("Rebalance Signal", "🔔 YES" if rebalance else "No")

st.markdown("---")

# ── Row 2: Gauge + Score History ─────────────────────────────
col_gauge, col_hist = st.columns([1, 2])

with col_gauge:
    st.subheader("Composite Score Gauge")
    fig_gauge = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=composite_score,
        number={"valueformat":".4f","font":{"size":28}},
        delta={"reference": snapshot.get("cycle_result",{}).get("prior_score") or composite_score,
               "valueformat":"+.4f"},
        gauge={
            "axis": {"range":[0,1],"tickwidth":2},
            "bar":  {"color": phase_colour},
            "steps": [
                {"range":[0.00,0.20],"color":"#fadbd8"},
                {"range":[0.20,0.40],"color":"#e8daef"},
                {"range":[0.40,0.60],"color":"#fef9e7"},
                {"range":[0.60,0.80],"color":"#d5f5e3"},
                {"range":[0.80,1.00],"color":"#a9dfbf"},
            ],
            "threshold": {"line":{"color":"red","width":4},"thickness":0.75,"value":composite_score},
        },
        title={"text": f"Phase: {phase}", "font":{"size":16}},
    ))
    fig_gauge.update_layout(height=300, margin=dict(l=20,r=20,t=40,b=20))
    st.plotly_chart(fig_gauge, use_container_width=True)

    # Phase bands legend
    bands = [("0.80–1.00","Strong Recovery","#1a7f37"),
             ("0.60–0.80","Early Expansion","#2d8b56"),
             ("0.40–0.60","Mid Cycle","#b7950b"),
             ("0.20–0.40","Late Cycle","#8e44ad"),
             ("0.00–0.20","Contraction","#c0392b")]
    for rng, lbl, col in bands:
        marker = "◀" if lbl == phase else "  "
        st.markdown(f"""<span style="color:{col};font-size:12px;">{marker} {rng} — {lbl}</span>""",
                    unsafe_allow_html=True)

with col_hist:
    st.subheader("Score History")
    if not hist_df.empty:
        fig_hist = go.Figure()
        fig_hist.add_trace(go.Scatter(
            x=hist_df["date"], y=hist_df["composite_score"].astype(float),
            mode="lines+markers+text", name="Score",
            text=hist_df["phase"], textposition="top center",
            line=dict(color=phase_colour, width=2),
            marker=dict(size=8),
        ))
        # Phase bands
        for lo, hi, col in [(0.8,1.0,"#a9dfbf"),(0.6,0.8,"#d5f5e3"),
                            (0.4,0.6,"#fef9e7"),(0.2,0.4,"#e8daef"),(0.0,0.2,"#fadbd8")]:
            fig_hist.add_hrect(y0=lo, y1=hi, fillcolor=col, opacity=0.3, line_width=0)
        fig_hist.update_layout(height=300, yaxis_range=[0,1],
                               margin=dict(l=20,r=20,t=20,b=20),
                               yaxis_title="Composite Score",
                               hovermode="x unified")
        st.plotly_chart(fig_hist, use_container_width=True)
    else:
        st.info("No history yet. Run the system at least once to build history.")

st.markdown("---")

# ── Row 3: Input Scorecard ────────────────────────────────────
st.subheader("📊 Full Input Scorecard (15 Indicators)")

ORDER = [
    "nifty_6m_change","pmi_manufacturing","repo_rate_trend","credit_growth","housing_starts",
    "gdp_growth","iip_growth","earnings_growth","auto_sales","gst_collections",
    "cpi_inflation","unemployment_rate","bank_npa_ratio","current_account_deficit","wpi_inflation",
]
LABELS = {
    "nifty_6m_change":"Nifty 50 6M Change","pmi_manufacturing":"PMI Manufacturing",
    "repo_rate_trend":"Repo Rate Trend (3M)","credit_growth":"Credit Growth YoY",
    "housing_starts":"Housing Starts","gdp_growth":"GDP Growth QoQ Ann.",
    "iip_growth":"IIP Industrial Prod.","earnings_growth":"Earnings Growth (Nifty)",
    "auto_sales":"Auto Sales YoY","gst_collections":"GST Collections YoY",
    "cpi_inflation":"CPI Inflation","unemployment_rate":"Unemployment Rate",
    "bank_npa_ratio":"Bank NPA Ratio","current_account_deficit":"Current A/C Deficit",
    "wpi_inflation":"WPI Inflation",
}
GROUPS_MAP = {k: "Leading" for k in ORDER[:5]}
GROUPS_MAP.update({k: "Coincident" for k in ORDER[5:10]})
GROUPS_MAP.update({k: "Lagging" for k in ORDER[10:]})

rows = []
for key in ORDER:
    ind = indicators.get(key, {})
    rows.append({
        "Group":     GROUPS_MAP.get(key,""),
        "Indicator": LABELS.get(key, key),
        "Value":     str(ind.get("value","—")) if ind.get("value") is not None else "⚠ MISSING",
        "Source":    str(ind.get("source","—"))[:50],
        "Data Date": str(ind.get("data_date","—"))[:20],
        "Status":    ind.get("fetch_status","—"),
        "Notes":     str(ind.get("notes",""))[:50],
    })

df_ind = pd.DataFrame(rows)

def colour_status(val):
    if val == "OK":              return "background-color: #d4edda"
    if val == "ESTIMATED":       return "background-color: #fff3cd"
    if val == "MANUAL_REQUIRED": return "background-color: #f8d7da"
    return ""

st.dataframe(
    df_ind.style.applymap(colour_status, subset=["Status"]),
    use_container_width=True,
    height=420,
)

st.markdown("---")

# ── Row 4: Sector Strategy + ETF ─────────────────────────────
col_sec, col_etf = st.columns(2)

with col_sec:
    st.subheader(f"🏦 Sector Strategy — {phase}")
    strategy = SECTOR_STRATEGY.get(phase, {})
    if strategy:
        st.success("**OVERWEIGHT:** " + ", ".join(strategy.get("overweight",[])))
        st.warning("**NEUTRAL:** " + ", ".join(strategy.get("neutral",[])))
        st.error("**UNDERWEIGHT:** " + ", ".join(strategy.get("underweight",[])))
        st.info(f"**Action:** {strategy.get('action','')}")
        st.caption(f"Market outlook: {strategy.get('market_outlook','')} | Historical: {strategy.get('model_years','')}")

with col_etf:
    st.subheader("📈 ETF Recommendations")
    etf_data = snapshot.get("sector_recs", {}).get("etf_recs", {})
    if etf_data:
        etf_rows = []
        for ticker, info in etf_data.items():
            etf_rows.append({
                "ETF": info.get("name",""),
                "Ticker": ticker,
                "Stance": info.get("stance","N"),
                "Tag": info.get("tag",""),
                "Price": f"₹{info['price']:,.0f}" if info.get("price") else "N/A",
                "RSI": f"{info['rsi']:.0f}" if info.get("rsi") else "N/A",
                "4W Momentum": f"{info['momentum_4w']:+.1f}%" if info.get("momentum_4w") is not None else "N/A",
                "MA Signal": info.get("ma_signal","N/A"),
            })
        df_etf = pd.DataFrame(etf_rows)
        def colour_tag(val):
            if val == "BUY":       return "background-color:#d4edda;color:#1a7f37;font-weight:bold"
            if val == "WATCHLIST": return "background-color:#fff3cd;color:#b7950b;font-weight:bold"
            if val == "AVOID":     return "background-color:#f8d7da;color:#c0392b;font-weight:bold"
            return ""
        st.dataframe(df_etf.style.applymap(colour_tag, subset=["Tag"]),
                     use_container_width=True, height=280)
    else:
        st.info("Run the system to populate ETF data.")

st.markdown("---")

# ── Row 5: History table ──────────────────────────────────────
st.subheader("📅 Historical Weekly Comparison")
if not hist_df.empty:
    display_cols = ["date","composite_score","phase"] + [
        "gdp_growth","pmi_manufacturing","cpi_inflation",
        "credit_growth","repo_rate_trend","auto_sales","gst_collections"
    ]
    avail = [c for c in display_cols if c in hist_df.columns]
    st.dataframe(hist_df[avail].tail(8).sort_values("date",ascending=False),
                 use_container_width=True)
else:
    st.info("No history yet. History builds automatically after each weekly run.")

# ── Footer ────────────────────────────────────────────────────
st.markdown("---")
st.caption(
    "India Macro Cycle System | Data: MOSPI · RBI · Finance Ministry · SIAM/FADA · yfinance | "
    "Free official sources only | Not financial advice"
)
