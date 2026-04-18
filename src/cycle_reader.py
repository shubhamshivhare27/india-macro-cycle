# ============================================================
#  src/cycle_reader.py
#  Reads the OUTPUT cells from Excel after formulas recalculate.
#  Computes momentum vs prior week from history CSV.
#  Returns full cycle result dict.
# ============================================================

import re
import csv
import time
import logging
import datetime
from pathlib import Path

logger = logging.getLogger("macro_cycle")

OUTPUT_CELLS = {
    "composite_score": "E39",
    "phase_label":     "H49",
}

PHASE_CLEAN = {
    "strong recovery": "STRONG RECOVERY",
    "early expansion": "EARLY EXPANSION",
    "mid cycle":       "MID CYCLE",
    "late cycle":      "LATE CYCLE",
    "contraction":     "CONTRACTION",
}

def _parse_phase(label_str: str) -> str:
    """Extract clean phase name from Excel label string."""
    s = str(label_str).lower()
    for key, clean in PHASE_CLEAN.items():
        if key in s:
            return clean
    # Fallback: map score to phase
    return "UNKNOWN"

def _score_to_phase(score: float) -> str:
    if score >= 0.80: return "STRONG RECOVERY"
    if score >= 0.60: return "EARLY EXPANSION"
    if score >= 0.40: return "MID CYCLE"
    if score >= 0.20: return "LATE CYCLE"
    return "CONTRACTION"


def read_from_excel(excel_path: Path, sheet_name: str, retries: int = 3) -> dict:
    """
    Read composite score and phase label from Excel output cells.
    Uses data_only=True to get formula-computed values.
    Retries to handle Excel recalculation delay.
    """
    from openpyxl import load_workbook

    composite_score = None
    phase_label     = None

    for attempt in range(1, retries + 1):
        try:
            time.sleep(2)  # give Excel formulas time to settle
            wb = load_workbook(str(excel_path), data_only=True)
            ws = wb[sheet_name]

            raw_score = ws["E39"].value
            raw_phase = ws["H49"].value

            if raw_score is not None:
                composite_score = float(str(raw_score))
            if raw_phase is not None:
                phase_label = str(raw_phase).strip()

            wb.close()

            if composite_score is not None:
                break
            logger.warning(f"  Attempt {attempt}: E39 returned None — retrying...")
            time.sleep(3)

        except Exception as e:
            logger.error(f"  Attempt {attempt}: Excel read error: {e}")
            time.sleep(3)

    if composite_score is None:
        logger.error("  Could not read composite score from Excel after retries")
        composite_score = 0.0
        phase_label     = "UNKNOWN"

    phase = _parse_phase(phase_label) if phase_label else _score_to_phase(composite_score)
    logger.info(f"  Excel output: Score={composite_score:.4f} | Phase={phase}")
    return {"composite_score": composite_score, "phase": phase, "raw_phase_label": phase_label}


def load_prior_week(history_path: Path) -> dict | None:
    """Read last week's row from macro_history.csv."""
    if not history_path.exists():
        return None
    try:
        rows = []
        with open(history_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
        if rows:
            last = rows[-1]
            return {
                "date":            last.get("date", ""),
                "composite_score": float(last.get("composite_score", 0)),
                "phase":           last.get("phase", ""),
            }
    except Exception as e:
        logger.error(f"  Could not read prior week from history: {e}")
    return None


def compute_cycle_result(excel_path: Path, sheet_name: str,
                         history_path: Path, indicators: dict) -> dict:
    """
    Full cycle reading pipeline:
    1. Read score + phase from Excel
    2. Load prior week for momentum
    3. Compute confidence based on data quality
    4. Determine rebalance flag
    Returns full cycle_result dict.
    """
    # 1. Read from Excel
    excel_out = read_from_excel(excel_path, sheet_name)
    score     = excel_out["composite_score"]
    phase     = excel_out["phase"]

    # 2. Momentum vs prior
    prior = load_prior_week(history_path)
    if prior:
        prior_score = prior["composite_score"]
        delta       = round(score - prior_score, 4)
        momentum    = "Rising" if delta > 0.02 else "Falling" if delta < -0.02 else "Stable"
        prior_phase = prior["phase"]
    else:
        delta       = None
        momentum    = "First Run"
        prior_phase = None
        prior_score = None

    # 3. Confidence
    live    = sum(1 for v in indicators.values() if v.get("fetch_status") == "OK" and not v.get("is_estimated"))
    est     = sum(1 for v in indicators.values() if v.get("is_estimated"))
    manual  = sum(1 for v in indicators.values() if v.get("fetch_status") == "MANUAL_REQUIRED")
    if live >= 12:     confidence = "High"
    elif live >= 9:    confidence = "Medium"
    else:              confidence = "Low"

    # 4. Rebalance flag — trigger if phase changed OR score delta > 0.05
    phase_changed   = (prior_phase is not None and prior_phase != phase)
    big_score_move  = (delta is not None and abs(delta) > 0.05)
    rebalance_flag  = phase_changed or big_score_move

    result = {
        "composite_score":  score,
        "phase":            phase,
        "prior_score":      prior_score,
        "score_delta":      delta,
        "momentum":         momentum,
        "prior_phase":      prior_phase,
        "phase_changed":    phase_changed,
        "confidence":       confidence,
        "live_count":       live,
        "estimated_count":  est,
        "manual_count":     manual,
        "rebalance_flag":   rebalance_flag,
        "run_date":         datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
    }

    logger.info(f"  Cycle result: {phase} | Score: {score:.3f} | Momentum: {momentum} | Confidence: {confidence}")
    return result


def print_cycle_result(result: dict) -> None:
    from colorama import Fore, Style

    PHASE_COLOURS = {
        "STRONG RECOVERY": Fore.GREEN + Style.BRIGHT,
        "EARLY EXPANSION": Fore.GREEN,
        "MID CYCLE":       Fore.YELLOW,
        "LATE CYCLE":      Fore.MAGENTA,
        "CONTRACTION":     Fore.RED,
    }
    colour = PHASE_COLOURS.get(result["phase"], Fore.WHITE)
    W = 70

    print()
    print(Fore.CYAN + Style.BRIGHT + "═" * W + Style.RESET_ALL)
    print(colour + f"  CYCLE PHASE:  {result['phase']}" + Style.RESET_ALL)
    score     = result["composite_score"]
    momentum  = result["momentum"]
    delta_str = f"{result['score_delta']:+.4f}" if result["score_delta"] is not None else "N/A"
    print(f"  Composite Score : {score:.4f}  ({score*100:.1f}% of maximum)")
    print(f"  Momentum        : {momentum}  (Δ {delta_str} vs prior week)")
    print(f"  Confidence      : {result['confidence']}  "
          f"({result['live_count']} live | {result['estimated_count']} estimated | {result['manual_count']} manual)")
    if result["phase_changed"]:
        print(Fore.YELLOW + f"  ⚠  PHASE CHANGED: {result['prior_phase']} → {result['phase']}" + Style.RESET_ALL)
    if result["rebalance_flag"]:
        print(Fore.YELLOW + "  🔔 REBALANCE SIGNAL TRIGGERED" + Style.RESET_ALL)
    print(Fore.CYAN + Style.BRIGHT + "═" * W + Style.RESET_ALL)
    print()
