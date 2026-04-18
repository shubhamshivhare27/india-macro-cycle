# ============================================================
#  src/excel_updater.py
#  Writes confirmed indicator values to Excel input cells only.
#  Preserves ALL formulas. Creates backup before writing.
#  Attempts live visual refresh via win32com on Windows.
# ============================================================

import re
import time
import logging
import shutil
import datetime
from pathlib import Path

logger = logging.getLogger("macro_cycle")

# ── Exact cell map (confirmed from Excel audit) ──────────────
CELL_MAP = {
    "nifty_6m_change":         "B8",
    "pmi_manufacturing":       "B9",
    "repo_rate_trend":         "B10",
    "credit_growth":           "B11",
    "housing_starts":          "B12",
    "gdp_growth":              "B16",
    "iip_growth":              "B17",
    "earnings_growth":         "B18",
    "auto_sales":              "B19",
    "gst_collections":         "B20",
    "cpi_inflation":           "B24",
    "unemployment_rate":       "B25",
    "bank_npa_ratio":          "B26",
    "current_account_deficit": "B27",
    "wpi_inflation":           "B28",
}

# ── Excel-format value coercion ──────────────────────────────
def _coerce_value(key: str, raw_value) -> str:
    """
    Convert fetched value to the string format Excel expects in each cell.
    Cells B8–B28 store text values like '15.2%', '56.4', '-25 bps', 'Rising'.
    """
    val = str(raw_value).strip()

    # Housing starts — text only
    if key == "housing_starts":
        for word in ["Rising", "Stable", "Falling"]:
            if word.lower() in val.lower():
                return word
        return "Stable"

    # Repo rate trend — keep bps format
    if key == "repo_rate_trend":
        m = re.search(r"([+-]?\d+)\s*bps?", val, re.IGNORECASE)
        if m:
            bps = int(m.group(1))
            return f"{bps:+d} bps" if bps != 0 else "0 bps"
        # If numeric without unit
        m2 = re.search(r"([+-]?\d+\.?\d*)", val)
        if m2:
            return f"{int(float(m2.group(1))):+d} bps"
        return "0 bps"

    # Numeric % values — strip % if present, return as "7.2%" format
    if key in ("nifty_6m_change","credit_growth","gdp_growth","iip_growth",
               "earnings_growth","auto_sales","gst_collections","cpi_inflation",
               "unemployment_rate","bank_npa_ratio","current_account_deficit","wpi_inflation"):
        m = re.search(r"-?\d+\.?\d*", val)
        if m:
            num = float(m.group())
            return f"{num:.1f}%"
        return val

    # PMI — plain number
    if key == "pmi_manufacturing":
        m = re.search(r"\d+\.?\d*", val)
        return m.group() if m else val

    return val


def backup_excel(excel_path: Path) -> Path:
    """Create timestamped backup of Excel file before writing."""
    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    bk_name  = excel_path.stem + f"_backup_{ts}" + excel_path.suffix
    bk_path  = excel_path.parent / "macro-cycle-system" / "data" / bk_name
    bk_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(excel_path, bk_path)
    logger.info(f"  Backup created: {bk_path.name}")
    return bk_path


def write_to_excel_openpyxl(excel_path: Path, indicators: dict, sheet_name: str) -> dict:
    """
    Write values to Excel using openpyxl.
    data_only=False ensures formulas are preserved.
    Returns dict of {cell: (old_value, new_value)}.
    """
    from openpyxl import load_workbook

    # Load preserving ALL formulas
    wb = load_workbook(str(excel_path), data_only=False, keep_vba=False)
    ws = wb[sheet_name]

    write_log = {}
    errors     = []

    for key, cell_ref in CELL_MAP.items():
        ind = indicators.get(key)
        if not ind:
            logger.warning(f"  ⚠ No data for {key}, skipping cell {cell_ref}")
            continue
        if ind.get("fetch_status") == "MANUAL_REQUIRED" and ind.get("value") is None:
            logger.warning(f"  ⚠ {key} still missing — skipping {cell_ref}")
            continue

        raw_val    = ind.get("value")
        formatted  = _coerce_value(key, raw_val)
        old_val    = ws[cell_ref].value

        try:
            ws[cell_ref] = formatted
            write_log[cell_ref] = {"key": key, "old": old_val, "new": formatted}
            logger.info(f"  ✍  {cell_ref} [{key}]: '{old_val}' → '{formatted}'")
        except Exception as e:
            errors.append(f"{cell_ref}: {e}")
            logger.error(f"  ❌ Failed writing {cell_ref}: {e}")

    # Save
    wb.save(str(excel_path))
    wb.close()

    if errors:
        logger.error(f"  {len(errors)} write error(s): {errors}")
    else:
        logger.info(f"  ✅ {len(write_log)} cells written successfully")

    return write_log


def refresh_excel_visual(excel_path: Path, sheet_name: str) -> bool:
    """
    Attempt to refresh Excel visually using win32com (Windows only).
    Returns True if successful, False if win32com not available.
    """
    try:
        import win32com.client as win32
        from colorama import Fore, Style

        print(f"  Opening Excel for visual refresh...")
        xl = win32.Dispatch("Excel.Application")
        xl.Visible = True
        xl.DisplayAlerts = False

        # Check if already open
        wb_obj = None
        for wb_open in xl.Workbooks:
            if Path(wb_open.FullName).resolve() == excel_path.resolve():
                wb_obj = wb_open
                break

        if wb_obj is None:
            wb_obj = xl.Workbooks.Open(str(excel_path))

        # Activate the target sheet
        wb_obj.Sheets(sheet_name).Activate()

        # Force recalculation
        xl.CalculateFullRebuild()
        time.sleep(2)  # wait for formulas

        # Save
        wb_obj.Save()
        print(Fore.GREEN + f"  ✅ Excel updated and visible on screen. Formulas recalculated." + Style.RESET_ALL)
        logger.info("  win32com visual refresh: SUCCESS")
        return True

    except ImportError:
        logger.info("  win32com not available — Excel updated silently via openpyxl")
        return False
    except Exception as e:
        logger.warning(f"  win32com refresh failed: {e} — Excel was still saved by openpyxl")
        return False


def update_excel(excel_path: Path, indicators: dict, sheet_name: str) -> dict:
    """
    Full update pipeline:
    1. Backup Excel
    2. Write values via openpyxl
    3. Attempt visual refresh via win32com
    4. Log all writes
    Returns write_log dict.
    """
    from colorama import Fore, Style

    print(f"\n  {'─'*70}")
    print(f"  WRITING TO EXCEL: {excel_path.name}")
    print(f"  Sheet: {sheet_name}")
    print(f"  {'─'*70}")

    # 1. Backup
    print("  Creating backup...")
    try:
        backup_excel(excel_path)
    except Exception as e:
        logger.error(f"  Backup failed: {e}")
        print(Fore.YELLOW + f"  ⚠ Backup failed ({e}) — proceeding anyway" + Style.RESET_ALL)

    # 2. Write values
    print("  Writing values to input cells...")
    write_log = {}
    try:
        write_log = write_to_excel_openpyxl(excel_path, indicators, sheet_name)
    except Exception as e:
        logger.error(f"  CRITICAL: openpyxl write failed: {e}")
        print(Fore.RED + f"  ❌ Excel write failed: {e}" + Style.RESET_ALL)
        raise

    # 3. Show write summary
    print()
    print(f"  {'Cell':<6} {'Indicator':<33} {'Old Value':<16} {'New Value':<16}")
    print(f"  {'─'*6} {'─'*33} {'─'*16} {'─'*16}")
    for cell, info in write_log.items():
        old = str(info["old"])[:15] if info["old"] else "—"
        new = str(info["new"])[:15]
        print(f"  {cell:<6} {info['key']:<33} {old:<16} {Fore.GREEN}{new:<16}{Style.RESET_ALL}")
    print()
    print(Fore.GREEN + f"  ✅ {len(write_log)}/15 cells written to Excel." + Style.RESET_ALL)

    # 4. Visual refresh
    print()
    refresh_excel_visual(excel_path, sheet_name)

    return write_log
